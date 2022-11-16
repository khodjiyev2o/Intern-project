from sqlalchemy.ext.asyncio import AsyncSession, async_object_session
from models.companies import Company, Member, Request, Quiz, Result
from schemas.companies import CompanyCreateSchema, CompanySchema, CompanyAlterSchema, RequestSchema, MemberSchema, QuizCreateSchema, QuizSchema, QuizAlterSchema, ResultSchema, QuizAnswerSchema, AverageScoreSchema, LastTimeQuiz, AverageScoreUserSchema
from sqlalchemy.future import select
from sqlalchemy import desc
from fastapi import HTTPException, Depends
from models.users import User
from services.users import get_user
from db import get_session, redis
from fastapi_pagination import Params
from fastapi_pagination.ext.async_sqlalchemy import paginate
from sqlalchemy.orm import selectinload
from datetime import datetime
from pickle import dumps, loads
from io import StringIO
from csv import writer, QUOTE_NONNUMERIC
from typing import Iterator, Optional, BinaryIO
from openpyxl import load_workbook
from io import BytesIO


class CompanyCRUD:
    def __init__(self, session: Optional[AsyncSession] = None, company: Optional[Company] = None):
        if not session:
            self.session = async_object_session(company)
        else:
            self.session = session
        self.company = company

    async def create_company(self, company: CompanyCreateSchema, user: User) -> CompanySchema:
        db_company = await self.session.execute(select(Company).filter(Company.name == company.name))
        db_company = db_company.scalars().first()
        if db_company:
            raise HTTPException(404, 'name already in use')
        else:
            new_company = Company(name=company.name, owner_id=user.id, description=company.description, visible=company.visible)
            self.session.add(new_company)
            await self.session.commit()
            return CompanySchema(id=new_company.id, name=new_company.name, owner=user.username, description=new_company.description, visible=new_company.visible)

    async def patch_company(self, company: CompanyAlterSchema, user: User) -> CompanySchema:
        if self.company.owner_id == user.id:
            if company.name:
                self.company.name = company.name
            if company.description:
                self.company.description = company.description
            if company.visible:
                self.company.visible = company.visible
            await self.session.commit()
            return CompanySchema(id=self.company.id, name=self.company.name, owner=user.username, description=self.company.description, visible=self.company.visible)
        else:
            raise HTTPException(404, 'no access')

    async def delete_company(self, user: User) -> CompanySchema:
        if self.company.owner_id == user.id:
            await self.session.delete(self.company)
            await self.session.commit()
            return CompanySchema(id=self.company.id, name=self.company.name, owner=user.username, description=self.company.description, visible=self.company.visible)
        else:
            raise HTTPException(404, 'no access')

    async def get_companies(self, page: int) -> list[CompanySchema]:
        params = Params(page=page, size=10)
        companies = await paginate(self.session, select(Company).options(selectinload(Company.owner)).filter(Company.visible == True), params=params)
        return [CompanySchema(id=company.id, name=company.name, owner=company.owner.username, description=company.description, visible=company.visible) for company in companies.items]

    async def get_company(self, id: int, user: User) -> CompanySchema:
        company = await self.session.get(Company, id)
        if company:
            if company.visible:
                return CompanySchema(id=company.id, name=company.name, owner=user.username, description=company.description, visible=company.visible)
            else:
                if company.owner_id == user.id:
                    return CompanySchema(id=company.id, name=company.name, owner=user.username, description=company.description, visible=company.visible)
                else:
                    raise HTTPException(404, 'company hidden')
        else:
            raise HTTPException(404, 'company not found')


async def get_company(id: int, session: AsyncSession = Depends(get_session), user: User = Depends(get_user)) -> Company:
    company = await session.get(Company, id)
    if company:
        if company.visible:
            return company
        else:
            if company.owner_id == user.id:
                return company
            else:
                raise HTTPException(404, 'company hidden')
    else:
        raise HTTPException(404, 'company not found')


class RequestCRUD:
    def __init__(self, session: Optional[AsyncSession] = None, user: Optional[User] = None, company: Optional[Company] = None):
        self.user = user
        self.company = company
        if user:
            self.session = async_object_session(user)
        elif company:
            self.session = async_object_session(company)
        else:
            self.session = session

    def side_bool_to_str(self, side) -> str:
        if side:
            return 'User requests access to company'
        else:
            return 'Company invites user'

    async def get_requests(self,  page: int, cur_user: Optional[User] = None, user_id: Optional[int] = None, company_id: Optional[Company] = None) -> list[RequestSchema]:
        params = Params(page=page, size=10)
        if user_id:
            requests = await paginate(self.session, select(Request).options(selectinload(Request.user)).options(selectinload(Request.company)).filter(Request.user_id == user_id), params=params)
        elif company_id:
            company = await self.session.get(Company, company_id)
            if company.owner_id == cur_user.id:
                requests = await paginate(self.session, select(Request).options(selectinload(Request.user)).options(selectinload(Request.company)).filter(Request.company_id == company_id), params=params)
            else:
                raise HTTPException(404, "you can't see other's company requests")
        return [RequestSchema(id=request.id, user=request.user.username, company=request.company.name, side=self.side_bool_to_str(request.side)) for request in requests.items]

    async def create_request(self, user_id: int, company_id: int, side: bool) -> RequestSchema:
        db_member = await self.session.get(Member, {'company_id': company_id, 'user_id': user_id})
        if db_member:
            raise HTTPException(404, 'already member of company')
        db_request = await self.session.execute(select(Request).filter(Request.company_id == company_id, Request.user_id == user_id))
        db_request = db_request.scalars().first()
        if db_request:
            raise HTTPException(404, 'request already submitted')

        user = await self.session.get(User, user_id)
        if not user:
            raise HTTPException(404, 'user not found')

        company = await self.session.get(Company, company_id)
        if not company:
            raise HTTPException(404, 'company not found')
        if user.id == company.owner_id:
            raise HTTPException(404, "can't invite yourself")

        request = Request(user_id=user_id, company_id=company_id, side=side)
        self.session.add(request)
        await self.session.commit()
        return RequestSchema(id=request.id, user=user.username, company=company.name, side=self.side_bool_to_str(request.side))


class MemberCRUD:
    def __init__(self, session: Optional[AsyncSession] = None, user: Optional[User] = None, company: Optional[Company] = None) -> None:
        self.user = user
        self.company = company
        if user:
            self.session = async_object_session(user)
        elif company:
            self.session = async_object_session(company)
        else:
            self.session = session

    async def remove_user(self, user_id: int, cur_user: User) -> MemberSchema:
        if self.company.owner_id == cur_user.id:
            member = await self.session.get(Member, {'company_id': self.company.id, 'user_id': user_id})
            if member:
                await self.session.delete(member)
                await self.session.commit()
                return MemberSchema(company=self.company.name, user=(await self.session.get(User, user_id)).username, admin=member.admin)
            else:
                raise HTTPException(404, 'member not found')
        else:
            raise HTTPException(404, "you can't remove users from others company")

    async def change_admin(self, user_id: int, cur_user: User, admin: bool) -> MemberSchema:
        if self.company.owner_id == cur_user.id:
            member = await self.session.get(Member, {'company_id': self.company.id, 'user_id': user_id})
            if member:
                member.admin = admin
                await self.session.commit()
                return MemberSchema(company=self.company.name, user=(await self.session.get(User, user_id)).username, admin=member.admin)
            else:
                raise HTTPException(404, 'member not found')
        else:
            raise HTTPException(404, "you can't change admin from others company")

    async def get_members(self, cur_user: User, page: int) -> list[MemberSchema]:
        params = Params(page=page, size=10)
        if self.company.owner_id == cur_user.id:
            members = await paginate(self.session, select(Member).filter(Member.company_id == self.company.id), params=params)
            return [MemberSchema(company=self.company.name, user=(await self.session.get(User, member.user_id)).username, admin=member.admin) for member in members.items]
        else:
            raise HTTPException(404, "you can't see other's companies members")

    async def review_request(self, request_id: int, response: str) -> MemberSchema:
        request = await self.session.get(Request, request_id)
        if not request:
            raise HTTPException(404, 'request not found')
        else:
            company = await self.session.get(Company, request.company_id)
            if response == 'decline':
                if request.side == True:
                    if self.user.id == company.owner_id:
                        await self.session.delete(request)
                        await self.session.commit()
                        raise HTTPException(404, 'request declined')
                    else:
                        raise HTTPException(404, "you can't review other's requests")
                elif request.side == False:
                    if self.user.id == request.user_id:
                        await self.session.delete(request)
                        await self.session.commit()
                        raise HTTPException(404, 'request declined')
                    else:
                        raise HTTPException(404, "you can't review other's requests")
            elif response == 'accept':
                if request.side == True:
                    if self.user.id == company.owner_id:
                        member = Member(company_id=request.company_id, user_id=request.user_id)
                        self.session.add(member)
                        await self.session.delete(request)
                        await self.session.commit()
                        user = await self.session.get(User, request.user_id)
                        return MemberSchema(company=company.name, user=user.username, admin=False)
                    else:
                        raise HTTPException(404, "you can't review other's requests")
                elif request.side == False:
                    if self.user.id == request.user_id:
                        member = Member(company_id=request.company_id, user_id=request.user_id)
                        self.session.add(member)
                        await self.session.delete(request)
                        await self.session.commit()
                        return MemberSchema(company=company.name, user=self.user.username, admin=False)
                    else:
                        raise HTTPException(404, "you can't review other's requests")
            else:
                raise HTTPException(404, 'invalid response')


class QuizCRUD:
    def __init__(self, session: Optional[AsyncSession] = None) -> None:
        self.session = session

    async def create_quiz(self, company: Company, user: User, quiz: QuizCreateSchema) -> QuizSchema:
        if company.owner_id != user.id:
            member = await self.session.get(Member, {'company_id': company.id, 'user_id': user.id})
            if not member or not member.admin:
                raise HTTPException(404, "you can't add quizzes to other's companies")
        db_quiz = await self.session.execute(select(Quiz).filter(Quiz.name == quiz.name))
        db_quiz = db_quiz.scalars().first()
        if db_quiz:
            raise HTTPException(404, 'Quiz already in db')
        db_quiz = Quiz(name=quiz.name, description=quiz.description, frequency=quiz.frequency, quiz={'questions': quiz.questions, 'answer_options': quiz.answer_options, 'correct_answers': quiz.correct_answers}, company_id=company.id)
        self.session.add(db_quiz)
        await self.session.commit()
        return QuizSchema(id=db_quiz.id, name=db_quiz.name, description=db_quiz.description, frequency=db_quiz.frequency, quiz=db_quiz.quiz, company_id=company.id)

    async def create_or_update_quiz_excel(self, company: Company, user: User, file: BinaryIO) -> QuizSchema:
        if company.owner_id != user.id:
            member = await self.session.get(Member, {'company_id': company.id, 'user_id': user.id})
            if not member or not member.admin:
                raise HTTPException(404, "you can't add or edit quizzes to other's companies")
        try:
            quiz_excel = load_workbook(BytesIO(file))
            quiz_excel = list(quiz_excel[quiz_excel.sheetnames[0]])
            name = quiz_excel[0][0].value
            description = quiz_excel[1][0].value
            frequency = int(quiz_excel[2][0].value)
            questions = []
            answer_options = []
            correct_answers = []
            for question in quiz_excel[3]:
                questions.append(question.value)
            for option_list in quiz_excel[4:-1]:
                options = []
                for option in option_list:
                    options.append(option.value)
                answer_options.append(options)
            for answer in quiz_excel[-1]:
                correct_answers.append(int(answer.value))
            quiz = QuizCreateSchema(name=name, description=description, frequency=frequency, questions=questions, answer_options=answer_options, correct_answers=correct_answers)
        except:
            raise HTTPException('Invalid excel file!')
        db_quiz = await self.session.execute(select(Quiz).filter(Quiz.name == name))
        db_quiz = db_quiz.scalars().first()
        if db_quiz:
            db_quiz.description = description
            db_quiz.frequency = frequency
            db_quiz.quiz = {'questions': questions, 'answer_options': answer_options, 'correct_answers': correct_answers}
            await self.session.commit()
        else:
            db_quiz = Quiz(name=name, description=description, frequency=frequency, quiz={'questions': questions, 'answer_options': answer_options, 'correct_answers': correct_answers}, company_id=company.id)
            self.session.add(db_quiz)
            await self.session.commit()
        return QuizSchema(id=db_quiz.id, name=db_quiz.name, description=db_quiz.description, frequency=db_quiz.frequency, quiz=db_quiz.quiz, company_id=company.id)

    async def patch_quiz(self, company: Company, user: User, quiz: QuizAlterSchema, quiz_id: int) -> QuizSchema:
        if company.owner_id != user.id:
            member = await self.session.get(Member, {'company_id': company.id, 'user_id': user.id})
            if not member or not member.admin:
                raise HTTPException(404, "you can't edit other's quizzes")

        db_quiz = await self.session.get(Quiz, quiz_id)
        if quiz.name:
            db_quiz.name = quiz.name
        if quiz.description:
            db_quiz.description = quiz.description
        if quiz.frequency:
            db_quiz.frequency = quiz.frequency
        if quiz.questions:
            quiz = {'questions': quiz.questions, 'answer_options': quiz.answer_options, 'correct_answers': quiz.correct_answers}
            db_quiz.quiz = quiz
        await self.session.commit()
        return QuizSchema(id=db_quiz.id, name=db_quiz.name, description=db_quiz.description, frequency=db_quiz.frequency, quiz=db_quiz.quiz, company_id=company.id)

    async def quizzes(self, company: Company, user: User, page: int) -> list[QuizSchema]:
        if company.owner_id != user.id:
            member = await self.session.get(Member, {'company_id': company.id, 'user_id': user.id})
            if not member:
                raise HTTPException(404, "you can't see other's quizzes")
        params = Params(page=page, size=10)
        quizzes = await paginate(self.session, select(Quiz).filter(Quiz.company_id == company.id), params=params)
        return [QuizSchema(id=db_quiz.id, name=db_quiz.name, description=db_quiz.description, frequency=db_quiz.frequency, quiz=db_quiz.quiz, company_id=company.id) for db_quiz in quizzes.items]

    async def delete_quiz(self, company: Company, user: User, quiz_id: int) -> QuizSchema:
        if company.owner_id != user.id:
            member = await self.session.get(Member, {'company_id': company.id, 'user_id': user.id})
            if not member or not member.admin:
                raise HTTPException(404, "you can't delete other's quizzes")
        db_quiz = await self.session.get(Quiz, quiz_id)
        if db_quiz:
            await self.session.delete(db_quiz)
            await self.session.commit()
            return QuizSchema(id=db_quiz.id, name=db_quiz.name, description=db_quiz.description, frequency=db_quiz.frequency, quiz=db_quiz.quiz, company_id=company.id)
        else:
            raise HTTPException(404, 'quiz not found')

    async def take_quiz(self, answers: QuizAnswerSchema, user: User, quiz_id: int) -> ResultSchema:
        quiz = await self.session.get(Quiz, quiz_id)
        company = await self.session.get(Company, quiz.company_id)
        if not quiz:
            raise HTTPException(404, 'quiz not found')
        member = await self.session.execute(select(Member).filter(Member.company_id == quiz.company_id, Member.user_id == user.id))
        member = member.scalars().first()
        if not member:
            raise HTTPException(404, 'access to quiz denied')
        correct = quiz.quiz['correct_answers']
        given = answers.answers
        if len(correct) != len(given):
            raise HTTPException(404, 'invalid answers')
        overall_questions = 0
        correct_answers = 0
        for answer, correct in zip(given, correct):
            if answer == 0:
                continue
            overall_questions += 1
            if answer == correct:
                correct_answers += 1

        result = Result(user_id=user.id, quiz_id=quiz_id, company_id=company.id, overall_questions=overall_questions, correct_answers=correct_answers)
        self.session.add(result)
        await self.session.commit()
        await redis.set(f'{user.id}-{quiz_id}-{company.id}-{datetime.utcnow()}', dumps(given), ex=48*3600)
        return ResultSchema(id=result.id, user_id=user.id, quiz_id=quiz_id, company_id=company.id, overall_questions=overall_questions, correct_answers=correct_answers)

    async def dump_results_user(self, user_id: int) -> Iterator:
        results = await self.session.execute(select(Result).filter(Result.user_id == user_id))
        results = results.scalars().all()
        csv = StringIO()
        write = writer(csv, quoting=QUOTE_NONNUMERIC)
        csvheader = ['user_id', 'quiz_id', 'overall_questions', 'correct_answers']
        write.writerow(csvheader)
        for result in results:
            write.writerow([result.user_id, result.quiz_id, result.overall_questions, result.correct_answers])
        return iter(csv.getvalue())

    async def dump_answers_user(self, user_id: int) -> Iterator:
        db_keys = redis.scan_iter(f'{user_id}-*')
        answers = []
        keys = []
        async for key in db_keys:
            answers.append(loads(await redis.get(key)))
            keys.append(key)
        csv = StringIO()
        write = writer(csv, quoting=QUOTE_NONNUMERIC)
        csvheader = ['user_id', 'quiz_id', 'answers']
        write.writerow(csvheader)
        for key, answer in zip(keys, answers):
            key = key.decode('utf-8').split('-')
            write.writerow([key[0], key[1], answer])
        return iter(csv.getvalue())

    async def dump_results_company(self, user: User, company: Company, user_id: Optional[int] = None, quiz_id: Optional[int] = None) -> Iterator:
        if company.owner_id != user.id:
            member = await self.session.get(Member, {'company_id': company.id, 'user_id': user.id})
            if not member or not member.admin:
                raise HTTPException(404, "you can't dump other's results")
        if user_id and quiz_id:
            results = await self.session.execute(select(Result).filter(Result.user_id == user_id, Result.quiz_id == quiz_id, Result.company_id == company.id))
        elif user_id:
            results = await self.session.execute(select(Result).filter(Result.user_id == user_id, Result.company_id == company.id))
        elif quiz_id:
            results = await self.session.execute(select(Result).filter(Result.quiz_id == quiz_id, Result.company_id == company.id))
        else:
            results = await self.session.execute(select(Result).filter(Result.company_id == company.id))
        results = results.scalars().all()
        csv = StringIO()
        write = writer(csv, quoting=QUOTE_NONNUMERIC)
        csvheader = ['user_id', 'quiz_id', 'overall_questions', 'correct_answers']
        write.writerow(csvheader)
        for result in results:
            write.writerow([result.user_id, result.quiz_id, result.overall_questions, result.correct_answers])
        return iter(csv.getvalue())

    async def dump_answers_company(self, user: User, company: Company, user_id: Optional[int] = None, quiz_id: Optional[int] = None) -> Iterator:
        if company.owner_id != user.id:
            member = await self.session.get(Member, {'company_id': company.id, 'user_id': user.id})
            if not member or not member.admin:
                raise HTTPException(404, "you can't dump other's results")
        if user_id and quiz_id:
            db_keys = redis.scan_iter(f'{user_id}-{quiz_id}-{company.id}-*')
        elif user_id:
            db_keys = redis.scan_iter(f'{user_id}-*-{company.id}-*')
        elif quiz_id:
            db_keys = redis.scan_iter(f'*-{quiz_id}-{company.id}-*')
        else:
            db_keys = redis.scan_iter(f'*-*-{company.id}-*')
        answers = []
        keys = []
        async for key in db_keys:
            answers.append(loads(await redis.get(key)))
            keys.append(key)
        csv = StringIO()
        write = writer(csv, quoting=QUOTE_NONNUMERIC)
        csvheader = ['user_id', 'quiz_id', 'answers']
        write.writerow(csvheader)
        for key, answer in zip(keys, answers):
            key = key.decode('utf-8').split('-')
            write.writerow([key[0], key[1], answer])
        return iter(csv.getvalue())

    async def average_score_company(self, user: User, company: Company, user_id: Optional[int]) -> list[AverageScoreSchema]:
        if company.owner_id != user.id:
            member = await self.session.get(Member, {'company_id': company.id, 'user_id': user.id})
            if not member or not member.admin:
                raise HTTPException(404, "you can't see other's results")
        if user_id:
            scores = await self.session.execute(select(Result).filter(Result.company_id == company.id, Result.user_id == user_id))
        else:
            scores = await self.session.execute(select(Result).filter(Result.company_id == company.id))
        scores = scores.scalars().all()
        if not scores:
            raise HTTPException(404, 'results not found')
        average_score = {}
        for score in scores:
            date = str(score.created_at.date())
            if date in average_score:
                overall, correct = average_score[date]
                average_score[date] = [overall+score.overall_questions, correct+score.correct_answers]
            else:
                average_score[date] = [score.overall_questions, score.correct_answers]
        return [AverageScoreSchema(date=date, average_score=average_score[date][1]/average_score[date][0]) for date in average_score]

    async def last_time_quiz(self, user: User, company: Company) -> list[LastTimeQuiz]:
        if company.owner_id != user.id:
            member = await self.session.get(Member, {'company_id': company.id, 'user_id': user.id})
            if not member or not member.admin:
                raise HTTPException(404, "you can't see other's results")
        members = await self.session.execute(select(Member).filter(Member.company_id == company.id))
        members = members.scalars().all()
        if not members:
            raise HTTPException(404, 'users not found')
        results = []
        for member in members:
            last_time = await self.session.execute(select(Result).filter(Result.user_id == member.user_id).order_by(desc(Result.created_at)))
            last_time = last_time.scalars().first()
            if not last_time:
                last_time = 'never'
            results.append(LastTimeQuiz(user_id=member.user_id, quiz_id=last_time.quiz_id, last_time=str(last_time.created_at.date())))
        return results

    async def average_score_user(self, user_id: int, quiz_id: Optional[int]) -> AverageScoreUserSchema:
        if not await self.session.get(User, user_id):
            raise HTTPException(404, 'user not found')
        if quiz_id:
            scores = await self.session.execute(select(Result).filter(Result.user_id == user_id, Result.quiz_id == quiz_id))
        else:
            scores = await self.session.execute(select(Result).filter(Result.user_id == user_id))
        scores = scores.scalars().all()
        if not scores:
            raise HTTPException(404, 'results not found')
        overall = 0
        correct = 0
        for score in scores:
            overall += score.overall_questions
            correct += score.correct_answers
        return AverageScoreUserSchema(user_id=user_id, average_score=correct/overall)

    async def last_time_quiz_user(self, user_id: int) -> list[LastTimeQuiz]:
        quizzes = await self.session.execute(select(Result).filter(Result.user_id == user_id).distinct(Result.quiz_id))
        quizzes = quizzes.scalars().all()
        if not quizzes:
            raise HTTPException(404, 'quizzes not found')
        results = []
        for quiz in quizzes:
            last_time = await self.session.execute(select(Result).filter(Result.user_id == quiz.user_id).order_by(desc(Result.created_at)))
            last_time = last_time.scalars().first()
            results.append(LastTimeQuiz(user_id=quiz.user_id, quiz_id=last_time.quiz_id, last_time=str(last_time.created_at.date())))
        return results
